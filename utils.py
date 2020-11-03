from django.contrib.auth.models import Permission
from django.http import HttpResponse
from openpyxl import Workbook
from exceptions_ import ImmutableUserError, UnmatchedPkError


def _group_has_perm(group, perm):
    try:
        group.permissions.get(codename=perm)
        return True
    except Permission.DoesNotExist:
        return False


def get_group_perms(group):
    return {
        'name': group.name,
        'add_user': _group_has_perm(group, 'add_user'),
        'update_user': _group_has_perm(group, 'update_user'),
        'delete_user': _group_has_perm(group, 'delete_user'),
        'set_password': _group_has_perm(group, 'set_password'),
        'manage_roles': _group_has_perm(group, 'manage_roles'),
        'view_user_account_audit_trail': _group_has_perm(group, 'view_user_account_audit_trail'),
        'finance_add_new': _group_has_perm(group, 'finance_add_new'),
        'finance_update': _group_has_perm(group, 'finance_update'),
        'finance_delete': _group_has_perm(group, 'finance_delete'),
        'finance_audit_trail': _group_has_perm(group, 'finance_audit_trail'),
        'finance_report': _group_has_perm(group, 'finance_report')
    }


def toggle_user_status(user, auth_user):
    """
    change user status to active or in-active
    :param user: user whose status is being toggled
    :param auth_user making the request
    :return:
    """
    user.is_active = False if user.is_active else True
    user.auth_user = auth_user
    user.save()


def check_is_immutable(user):
    """check if user account is immutable: cannot be changed or modified"""
    return user.is_immutable is True


def assert_immutable_user(user):
    if check_is_immutable(user):
        raise ImmutableUserError('Cannot modify or delete immutable user!')


def assert_unmatched_pk(pk1, pk2):
    try:
        assert int(pk1) == int(pk2)
    except AssertionError:
        raise UnmatchedPkError(f'{pk1} != {pk2}')


def dump_to_excel(query_set, columns, filename):
    """
    Dumps django model query set to excel using pandas
    """
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename={filename}.xlsx'
    workbook = Workbook()
    worksheet = workbook.active
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    for data in query_set:
        row_num += 1

        # Define the data for each cell in the row
        row = [data[i] for i in range(len(columns))]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)
    return response
